#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path


PROJECT_ROOT = Path(__file__).parent.parent
DEFAULT_SOURCE = PROJECT_ROOT.parent / "___Course_Offering_26-27_Student_Version_1.1.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "icai" / "extraccion_materias" / "icai_catalogo.csv"
SHEET_NAME = "OFFER 2026-2027 PROVISIONAL"

CSV_FIELDS = [
    "language", "term", "schedule", "studies", "degree", "ects", "ects_semester",
    "codigo", "nombre", "official_name", "english_name", "note_comment", "url_guia",
    "source_label", "program_label", "source_url", "source_file", "syllabus_links_json",
    "availability_label", "timing_risk_label", "level_label", "permission_label",
    "language_label", "schedule_label",
]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _number(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _term_and_language(value: str) -> tuple[str, str]:
    parts = [part.strip() for part in _clean(value).split(".", 1)]
    language = parts[0].title() if parts else ""
    raw_term = parts[1].lower() if len(parts) > 1 else ""
    if raw_term == "fall":
        return "Fall", language
    if raw_term in {"all-year", "full-year"}:
        return "All-year", language
    return "", language


def _studies(degree: str, year) -> str:
    degree_upper = degree.upper()
    if degree_upper in {"ADE", "BA", "BIO", "DCC", "DHP", "GITI", "GITT", "IMAT", "SAPIENS"}:
        return "Undergraduate"
    if degree_upper.startswith("M"):
        return "Master"
    numeric_year = _number(year)
    return "Undergraduate" if numeric_year is not None and numeric_year <= 4 else "Master"


def _permission(note: str, studies: str) -> str:
    if studies == "Master" or "permission" in note.lower():
        return "permission_required"
    return "standard"


def _normalized_name(official_name: str, english_name: str) -> str:
    if not official_name:
        return english_name
    if not english_name or official_name.casefold() == english_name.casefold():
        return official_name
    return f"{official_name} / {english_name}"


def _merge_group(courses: list[dict]) -> dict:
    first = courses[0]
    languages = sorted({course["language"] for course in courses if course["language"]})
    notes = list(dict.fromkeys(course["note_comment"] for course in courses if course["note_comment"]))
    language = " / ".join(languages)
    note = " | ".join(notes)
    studies = first["studies"]
    permission = _permission(note, studies)
    labels = {
        "availability": "exchange_term" if first["term"] == "Fall" else "exchange_term_all_year",
        "timing_risk": "low" if first["term"] == "Fall" else "medium",
        "level": "undergraduate" if studies == "Undergraduate" else "postgraduate",
        "permission": permission,
        "language": language.lower().replace(" / ", "_or_") if language else "unknown",
        "schedule": "unknown",
    }
    return {
        **first,
        "language": language,
        "note_comment": note,
        "labels": labels,
        "availability_label": labels["availability"],
        "timing_risk_label": labels["timing_risk"],
        "level_label": labels["level"],
        "permission_label": labels["permission"],
        "language_label": labels["language"],
        "schedule_label": labels["schedule"],
    }


def select_available_courses(rows: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        if _clean(row.get("visible")).lower() != "yes":
            continue
        term, language = _term_and_language(row.get("language. term", ""))
        if term not in {"Fall", "All-year"}:
            continue
        code = _clean(row.get("course code"))
        if not code:
            continue
        official_name = _clean(row.get("official name"))
        english_name = _clean(row.get("english name"))
        ects = _number(row.get("ects"))
        year_number = _number(row.get("year"))
        year = int(year_number) if year_number is not None and year_number.is_integer() else year_number
        degree_code = _clean(row.get("degree"))
        studies = _studies(degree_code, year)
        course = {
            "codigo": code,
            "nombre": _normalized_name(official_name, english_name),
            "official_name": official_name,
            "english_name": english_name,
            "institucion": "icai_comillas",
            "term": term,
            "studies": studies,
            "degree": f"{year}-{degree_code}" if year is not None else degree_code,
            "degree_code": degree_code,
            "year": year,
            "schedule": "Unknown",
            "language": language,
            "ects": ects,
            "ects_semester": ects / 2 if term == "All-year" and ects is not None else ects,
            "note_comment": _clean(row.get("note/comment")),
            "url_guia": "",
            "source_label": "course_offering_2026_2027_v1_1",
            "program_label": "sapiens" if degree_code.upper() == "SAPIENS" else "regular_icai",
            "source_url": "",
            "source_file": "___Course_Offering_26-27_Student_Version_1.1.xlsx",
            "syllabus_links": [],
            "guia": {},
        }
        grouped.setdefault(code, []).append(course)
    return [_merge_group(grouped[code]) for code in sorted(grouped)]


def load_excel_rows(path: Path, sheet_name: str = SHEET_NAME) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Falta openpyxl. Instálalo con: python3 -m pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja requerida: {sheet_name}")
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [_clean(value) for value in next(values)]
    return [dict(zip(headers, row)) for row in values]


def write_catalog_csv(path: Path, catalog: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for course in catalog:
            writer.writerow({
                **{field: course.get(field, "") for field in CSV_FIELDS},
                "syllabus_links_json": json.dumps(course.get("syllabus_links", []), ensure_ascii=False),
            })


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa el catálogo ICAI Fall/All-year desde el Excel oficial")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    catalog = select_available_courses(load_excel_rows(args.source))
    write_catalog_csv(args.output, catalog)
    fall = sum(course["term"] == "Fall" for course in catalog)
    all_year = sum(course["term"] == "All-year" for course in catalog)
    print(f"OK {args.output.name}: {len(catalog)} códigos únicos (Fall={fall}, All-year={all_year})")


if __name__ == "__main__":
    main()
